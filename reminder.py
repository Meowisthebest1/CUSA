# reminder.py
# Sends reminder emails 24h and 1h before reserved events.
# Run every 10 minutes via Task Scheduler / cron.

import os
import uuid
import smtplib
from datetime import datetime, timedelta
from email.message import EmailMessage
from urllib.parse import quote

import openpyxl

try:
    import tomllib  # py3.11+
except Exception:
    tomllib = None

SECRETS_PATH = os.path.join(".streamlit", "secrets.toml")

def load_secrets():
    # Minimal TOML parser using tomllib if available; else environment variables.
    secrets = {}
    if tomllib and os.path.exists(SECRETS_PATH):
        with open(SECRETS_PATH, "rb") as f:
            secrets = tomllib.load(f)
    def get(k, d=""):
        return secrets.get(k, os.getenv(k, d))
    return {
        "EXCEL_PATH": get("EXCEL_PATH", "Student Ambassador Sign Up Sheet.xlsx"),
        "SHEET_NAME": get("SHEET_NAME", "2025-2026"),
        "HEADER_ROW": int(get("HEADER_ROW", 3)),
        "SMTP_HOST": get("SMTP_HOST", ""),
        "SMTP_PORT": int(get("SMTP_PORT", 587)),
        "SMTP_USER": get("SMTP_USER", ""),
        "SMTP_PASS": get("SMTP_PASS", ""),
        "FROM_EMAIL": get("FROM_EMAIL", get("SMTP_USER", "")),
    }

def smtp_ready(cfg):
    return bool(cfg["SMTP_HOST"] and cfg["SMTP_USER"] and cfg["SMTP_PASS"] and cfg["FROM_EMAIL"])

def fmt_ics_dt(dt: datetime) -> str:
    return dt.strftime("%Y%m%dT%H%M%S")

def make_ics(event_title: str, start_dt: datetime, end_dt: datetime, location: str, description: str, uid: str) -> str:
    return "\n".join([
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//CUSA//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{fmt_ics_dt(datetime.now())}",
        f"DTSTART:{fmt_ics_dt(start_dt)}",
        f"DTEND:{fmt_ics_dt(end_dt)}",
        f"SUMMARY:{event_title}",
        f"LOCATION:{location}",
        f"DESCRIPTION:{description}",
        "END:VEVENT",
        "END:VCALENDAR",
        ""
    ])

def google_calendar_link(title: str, start_dt: datetime, end_dt: datetime, location: str, details: str) -> str:
    dates = f"{start_dt.strftime('%Y%m%dT%H%M%S')}/{end_dt.strftime('%Y%m%dT%H%M%S')}"
    return (
        "https://calendar.google.com/calendar/render?action=TEMPLATE"
        f"&text={quote(title)}"
        f"&dates={quote(dates)}"
        f"&location={quote(location)}"
        f"&details={quote(details)}"
    )

def send_email(cfg, to_email: str, subject: str, body: str, ics_text: str):
    msg = EmailMessage()
    msg["From"] = cfg["FROM_EMAIL"]
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body)
    msg.add_attachment(ics_text.encode("utf-8"), maintype="text", subtype="calendar", filename="event.ics")
    with smtplib.SMTP(cfg["SMTP_HOST"], cfg["SMTP_PORT"], timeout=30) as s:
        s.starttls()
        s.login(cfg["SMTP_USER"], cfg["SMTP_PASS"])
        s.send_message(msg)

TRACK_HEADERS = ["EMAIL", "USER_ID", "RESERVED_AT", "SENT_24H", "SENT_1H", "GCAL_UID"]

def get_headers(ws, header_row: int):
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().upper()] = col
    return headers

def ensure_tracking(ws, header_row: int):
    headers = get_headers(ws, header_row)
    col = 1
    while ws.cell(header_row, col).value is not None and col <= ws.max_column:
        col += 1
    for th in TRACK_HEADERS:
        if th.upper() not in headers:
            ws.cell(header_row, col).value = th
            col += 1

def parse_dt(dv, tv):
    if not dv or not tv:
        return None
    # dv may be date or datetime
    if isinstance(dv, datetime):
        d = dv.date()
    else:
        d = dv
    if isinstance(tv, datetime):
        t = tv.time()
    else:
        t = tv
    return datetime(d.year, d.month, d.day, t.hour, t.minute, t.second)

def main():
    cfg = load_secrets()
    if not smtp_ready(cfg):
        print("SMTP not configured. Edit .streamlit/secrets.toml.")
        return

    wb = openpyxl.load_workbook(cfg["EXCEL_PATH"])
    ws = wb[cfg["SHEET_NAME"]]
    ensure_tracking(ws, cfg["HEADER_ROW"])
    headers = get_headers(ws, cfg["HEADER_ROW"])

    def col(name): return headers[name.upper()]

    now = datetime.now()
    sent_any = 0

    for r in range(cfg["HEADER_ROW"] + 1, ws.max_row + 1):
        event = ws.cell(r, col("EVENT")).value
        dv = ws.cell(r, col("DATE")).value
        stv = ws.cell(r, col("START TIME")).value
        env = ws.cell(r, col("END TIME")).value
        email = str(ws.cell(r, col("EMAIL")).value or "").strip()

        if not (event and dv and stv and env and email):
            continue

        start_dt = parse_dt(dv, stv)
        end_dt = parse_dt(dv, env)
        if not start_dt or not end_dt:
            continue
        if end_dt <= start_dt:
            end_dt = end_dt + timedelta(days=1)

        sent24 = bool(ws.cell(r, col("SENT_24H")).value)
        sent1 = bool(ws.cell(r, col("SENT_1H")).value)

        delta = start_dt - now

        title = f"{event} (Student Ambassador)"
        location = str(ws.cell(r, col("LOCATION")).value or "").strip()
        contact = str(ws.cell(r, col("CONTACT PERSON")).value or "").strip()
        details = f"Reminder from CUSA. Contact: {contact}"
        gcal = google_calendar_link(title, start_dt, end_dt, location, details)
        uid = str(ws.cell(r, col("GCAL_UID")).value or "").strip() or str(uuid.uuid4())
        ws.cell(r, col("GCAL_UID")).value = uid
        ics = make_ics(title, start_dt, end_dt, location, details, uid)

        if timedelta(hours=23, minutes=30) <= delta <= timedelta(hours=24, minutes=30) and not sent24:
            body = f"Hi,\n\n⏰ Reminder: your event is in ~24 hours.\nEvent: {event}\nWhen: {start_dt:%A, %B %d, %Y at %I:%M %p}\nLocation: {location}\n\nGoogle Calendar link:\n{gcal}\n\n— CUSA"
            send_email(cfg, email, f"24-hour reminder: {event}", body, ics)
            ws.cell(r, col("SENT_24H")).value = True
            sent_any += 1

        if timedelta(minutes=30) <= delta <= timedelta(hours=1, minutes=30) and not sent1:
            body = f"Hi,\n\n⏰ Reminder: your event is in ~1 hour.\nEvent: {event}\nWhen: {start_dt:%A, %B %d, %Y at %I:%M %p}\nLocation: {location}\n\nGoogle Calendar link:\n{gcal}\n\n— CUSA"
            send_email(cfg, email, f"1-hour reminder: {event}", body, ics)
            ws.cell(r, col("SENT_1H")).value = True
            sent_any += 1

    wb.save(cfg["EXCEL_PATH"])
    print(f"Done. Sent {sent_any} reminders.")

if __name__ == "__main__":
    main()
