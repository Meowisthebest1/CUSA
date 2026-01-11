# CUSA.py
# Cleary University Student Ambassadors (CUSA)
# Streamlit: signup/login + Excel reservations + email confirmation + Google Calendar invite
# Forum w/ profile pics + admin moderation + admin management

from __future__ import annotations

import os
import re
import sqlite3
import uuid
import smtplib
from dataclasses import dataclass
from datetime import datetime, timedelta, date, time
from email.message import EmailMessage
from typing import Dict, List, Optional, Tuple
from urllib.parse import quote

import bcrypt
import openpyxl
import streamlit as st


# ----------------------------
# SAFE SECRETS (no crash)
# ----------------------------
def safe_secret(key: str, default=None):
    try:
        return st.secrets.get(key, default)  # works when secrets exist
    except Exception:
        return os.getenv(key, default)


# ----------------------------
# Config
# ----------------------------
DEFAULT_LOGO_URL = "https://www.internationalstudentinsurance.com/blog/2018/04/announcing-new-k-12-plan.html/cleary-university-seal-logo"

EXCEL_PATH = safe_secret("EXCEL_PATH", "Student Ambassador Sign Up Sheet.xlsx")
SHEET_NAME = safe_secret("SHEET_NAME", "2025-2026")
HEADER_ROW = int(safe_secret("HEADER_ROW", 3))
PROFILE_PIC_DIR = safe_secret("PROFILE_PIC_DIR", "profile_pics")
LOGO_PATH = safe_secret("LOGO_PATH", "Cleary-University-Seal-Logo.png")

SMTP_HOST = safe_secret("SMTP_HOST", "")
SMTP_PORT = int(safe_secret("SMTP_PORT", 587))
SMTP_USER = safe_secret("SMTP_USER", "")
SMTP_PASS = safe_secret("SMTP_PASS", "")
FROM_EMAIL = safe_secret("FROM_EMAIL", SMTP_USER)

BOOTSTRAP_ADMIN_EMAIL = safe_secret("BOOTSTRAP_ADMIN_EMAIL", "")
BOOTSTRAP_ADMIN_PASSWORD = safe_secret("BOOTSTRAP_ADMIN_PASSWORD", "")
BOOTSTRAP_ADMIN_FIRST = safe_secret("BOOTSTRAP_ADMIN_FIRST", "Admin")
BOOTSTRAP_ADMIN_LAST = safe_secret("BOOTSTRAP_ADMIN_LAST", "User")

FORCED_ADMIN_EMAILS = {"jhaley627@my.cleary.edu"}

USERS_DB = "users.db"
FORUM_DB = "forum.db"

TRACK_HEADERS = ["EMAIL", "USER_ID", "RESERVED_AT", "SENT_24H", "SENT_1H", "GCAL_UID"]


def smtp_ready() -> bool:
    return bool(SMTP_HOST and SMTP_USER and SMTP_PASS and FROM_EMAIL)


# ----------------------------
# UI (mobile-first, black w red/blue accents)
# ----------------------------
st.set_page_config(page_title="Cleary University Student Ambassadors", page_icon="🎓", layout="centered")

st.markdown("""
<style>
:root{
  --bg:#0b0b0f; --card:rgba(255,255,255,.06); --border:rgba(255,255,255,.12);
  --text:rgba(255,255,255,.92); --muted:rgba(255,255,255,.70);
}
html, body, [class*="stApp"]{ background:var(--bg)!important; color:var(--text)!important;}
.block-container{ padding-top:.8rem; max-width:760px;}
.card{ background:var(--card); border:1px solid var(--border); border-radius:18px; padding:14px; margin:10px 0;}
.pill{display:inline-block; padding:4px 10px; border-radius:999px; border:1px solid var(--border); font-size:.85rem; color:var(--muted)!important; margin-right:6px;}
div.stButton>button{ border-radius:14px; border:1px solid var(--border); background:rgba(255,255,255,.06); color:var(--text)!important;}
div.stButton>button[kind="primary"]{ background:linear-gradient(90deg, rgba(209,31,47,.85), rgba(43,108,255,.85)); border:1px solid rgba(255,255,255,.18); color:white!important;}
input, textarea{ background:rgba(255,255,255,.05)!important; color:var(--text)!important; border:1px solid var(--border)!important; border-radius:12px!important;}
</style>
""", unsafe_allow_html=True)

def resolve_logo():
    if LOGO_PATH and os.path.exists(LOGO_PATH):
        return LOGO_PATH
    # also try relative next to script
    p = os.path.join(os.path.dirname(__file__), "Cleary-University-Seal-Logo.png")
    if os.path.exists(p):
        return p
    return DEFAULT_LOGO_URL

st.markdown('<div style="text-align:center">', unsafe_allow_html=True)
st.image(resolve_logo(), width=120)
st.markdown('<div style="font-size:1.55rem;font-weight:800;letter-spacing:-.02em;margin:.35rem 0 .2rem 0;">Cleary University Student Ambassadors</div>', unsafe_allow_html=True)
st.markdown('<div style="color:rgba(255,255,255,.70);margin-bottom:.5rem;">Signups • Confirmations • Forum • Calendar invites</div>', unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

# If SMTP missing, show a very direct callout so you know it's not "broken"
if not smtp_ready():
    st.error(
        "EMAIL IS NOT CONFIGURED YET.\n\n"
        "Open `.streamlit/secrets.toml` and fill in SMTP_HOST/SMTP_USER/SMTP_PASS/FROM_EMAIL.\n"
        "After you save it, restart Streamlit."
    )


# ----------------------------
# Utilities
# ----------------------------
def sanitize_key(s: str) -> str:
    s = s.strip().lower()
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "user"

def ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)

def profile_pic_path(email: str) -> str:
    ensure_dir(PROFILE_PIC_DIR)
    return os.path.join(PROFILE_PIC_DIR, f"{sanitize_key(email)}.png")

def save_profile_pic(email: str, file_bytes: bytes) -> str:
    path = profile_pic_path(email)
    with open(path, "wb") as f:
        f.write(file_bytes)
    return path

def load_profile_pic(email: str) -> Optional[str]:
    path = profile_pic_path(email)
    return path if os.path.exists(path) else None


# ----------------------------
# Users DB
# ----------------------------
def users_conn():
    c = sqlite3.connect(USERS_DB, check_same_thread=False)
    c.execute("""
    CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        first_name TEXT NOT NULL,
        last_name TEXT NOT NULL,
        email TEXT NOT NULL UNIQUE,
        password_hash BLOB NOT NULL,
        is_admin INTEGER NOT NULL DEFAULT 0
    )
    """)
    c.commit()
    return c

def create_user(first_name: str, last_name: str, email: str, password: str, is_admin: bool=False) -> Tuple[bool, str]:
    if not first_name.strip() or not last_name.strip():
        return False, "First and last name are required."
    if "@" not in email or "." not in email:
        return False, "Please enter a valid email."
    if len(password) < 8:
        return False, "Password must be at least 8 characters."
    pw_hash = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt())
    c = users_conn()
    try:
        c.execute(
            "INSERT INTO users (first_name, last_name, email, password_hash, is_admin) VALUES (?,?,?,?,?)",
            (first_name.strip(), last_name.strip(), email.strip().lower(), pw_hash, 1 if is_admin else 0)
        )
        c.commit()
        return True, "Account created."
    except sqlite3.IntegrityError:
        return False, "That email is already registered."
    finally:
        c.close()

def authenticate(email: str, password: str) -> Optional[Dict]:
    email_n = email.strip().lower()
    c = users_conn()
    row = c.execute(
        "SELECT id, first_name, last_name, email, password_hash, is_admin FROM users WHERE email=?",
        (email_n,)
    ).fetchone()
    c.close()
    if not row:
        return None
    uid, fn, ln, em, pw_hash, is_admin = row
    if bcrypt.checkpw(password.encode("utf-8"), pw_hash):
        forced = email_n in FORCED_ADMIN_EMAILS
        return {"id": str(uid), "first_name": fn, "last_name": ln, "email": em, "is_admin": bool(is_admin) or forced}
    return None

def set_user_admin(email: str, is_admin: bool) -> Tuple[bool, str]:
    email_n = email.strip().lower()
    c = users_conn()
    row = c.execute("SELECT id FROM users WHERE email=?", (email_n,)).fetchone()
    if not row:
        c.close()
        return False, "User not found. They must sign up first."
    c.execute("UPDATE users SET is_admin=? WHERE email=?", (1 if is_admin else 0, email_n))
    c.commit()
    c.close()
    return True, "Updated."

def ensure_bootstrap_admin():
    if not (BOOTSTRAP_ADMIN_EMAIL and BOOTSTRAP_ADMIN_PASSWORD):
        return
    c = users_conn()
    row = c.execute("SELECT id FROM users WHERE email=?", (BOOTSTRAP_ADMIN_EMAIL.strip().lower(),)).fetchone()
    c.close()
    if row:
        return
    create_user(BOOTSTRAP_ADMIN_FIRST, BOOTSTRAP_ADMIN_LAST, BOOTSTRAP_ADMIN_EMAIL, BOOTSTRAP_ADMIN_PASSWORD, is_admin=True)

def ensure_forced_admins():
    c = users_conn()
    for em in FORCED_ADMIN_EMAILS:
        row = c.execute("SELECT id FROM users WHERE email=?", (em,)).fetchone()
        if row:
            c.execute("UPDATE users SET is_admin=1 WHERE email=?", (em,))
    c.commit()
    c.close()

ensure_bootstrap_admin()
ensure_forced_admins()


# ----------------------------
# Forum DB
# ----------------------------
def forum_conn():
    c = sqlite3.connect(FORUM_DB, check_same_thread=False)
    c.execute("""
    CREATE TABLE IF NOT EXISTS posts (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        author_email TEXT NOT NULL,
        author_name TEXT NOT NULL,
        title TEXT NOT NULL,
        body TEXT NOT NULL,
        locked INTEGER NOT NULL DEFAULT 0,
        deleted INTEGER NOT NULL DEFAULT 0
    )
    """)
    c.execute("""
    CREATE TABLE IF NOT EXISTS replies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        post_id INTEGER NOT NULL,
        created_at TEXT NOT NULL,
        author_email TEXT NOT NULL,
        author_name TEXT NOT NULL,
        body TEXT NOT NULL,
        deleted INTEGER NOT NULL DEFAULT 0,
        FOREIGN KEY(post_id) REFERENCES posts(id)
    )
    """)
    c.commit()
    return c

def create_post(author_email: str, author_name: str, title: str, body: str):
    c = forum_conn()
    c.execute(
        "INSERT INTO posts (created_at, author_email, author_name, title, body) VALUES (?,?,?,?,?)",
        (datetime.now().isoformat(timespec="seconds"), author_email, author_name, title.strip(), body.strip())
    )
    c.commit()
    c.close()

def list_posts(limit: int = 100, include_deleted: bool = False):
    c = forum_conn()
    if include_deleted:
        rows = c.execute(
            "SELECT id, created_at, author_email, author_name, title, body, locked, deleted FROM posts ORDER BY id DESC LIMIT ?",
            (limit,)
        ).fetchall()
    else:
        rows = c.execute(
            "SELECT id, created_at, author_email, author_name, title, body, locked, deleted FROM posts WHERE deleted=0 ORDER BY id DESC LIMIT ?",
            (limit,)
        ).fetchall()
    c.close()
    return rows

def create_reply(post_id: int, author_email: str, author_name: str, body: str):
    c = forum_conn()
    post = c.execute("SELECT locked, deleted FROM posts WHERE id=?", (post_id,)).fetchone()
    if not post or post[1] == 1:
        c.close()
        raise ValueError("Post not found.")
    if post[0] == 1:
        c.close()
        raise ValueError("Thread is locked.")
    c.execute(
        "INSERT INTO replies (post_id, created_at, author_email, author_name, body) VALUES (?,?,?,?,?)",
        (post_id, datetime.now().isoformat(timespec="seconds"), author_email, author_name, body.strip())
    )
    c.commit()
    c.close()

def list_replies(post_id: int, include_deleted: bool = False):
    c = forum_conn()
    if include_deleted:
        rows = c.execute(
            "SELECT id, created_at, author_email, author_name, body, deleted FROM replies WHERE post_id=? ORDER BY id ASC",
            (post_id,)
        ).fetchall()
    else:
        rows = c.execute(
            "SELECT id, created_at, author_email, author_name, body, deleted FROM replies WHERE post_id=? AND deleted=0 ORDER BY id ASC",
            (post_id,)
        ).fetchall()
    c.close()
    return rows

def set_lock(post_id: int, locked: bool):
    c = forum_conn()
    c.execute("UPDATE posts SET locked=? WHERE id=?", (1 if locked else 0, post_id))
    c.commit()
    c.close()

def soft_delete_post(post_id: int):
    c = forum_conn()
    c.execute("UPDATE posts SET deleted=1 WHERE id=?", (post_id,))
    c.execute("UPDATE replies SET deleted=1 WHERE post_id=?", (post_id,))
    c.commit()
    c.close()

def soft_delete_reply(reply_id: int):
    c = forum_conn()
    c.execute("UPDATE replies SET deleted=1 WHERE id=?", (reply_id,))
    c.commit()
    c.close()


# ----------------------------
# Excel helpers
# ----------------------------
@dataclass
class Slot:
    row: int
    event: str
    location: str
    start_dt: datetime
    end_dt: datetime
    hours: float
    contact: str
    first_name: str
    last_name: str
    completed: str
    email: str

def load_wb(path: str):
    return openpyxl.load_workbook(path)

def get_headers(ws, header_row: int) -> Dict[str, int]:
    headers = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(header_row, col).value
        if isinstance(v, str) and v.strip():
            headers[v.strip().upper()] = col
    return headers

def ensure_tracking_columns(path: str, sheet_name: str, header_row: int) -> None:
    wb = load_wb(path)
    ws = wb[sheet_name]
    headers = get_headers(ws, header_row)

    col = 1
    while ws.cell(header_row, col).value is not None and col <= ws.max_column:
        col += 1

    for th in TRACK_HEADERS:
        if th.upper() not in headers:
            ws.cell(header_row, col).value = th
            col += 1

    wb.save(path)

def parse_excel_date(v) -> Optional[date]:
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None

def parse_excel_time(v) -> Optional[time]:
    if isinstance(v, datetime):
        return v.time()
    if isinstance(v, time):
        return v
    return None

def slot_datetime(d: date, t: time) -> datetime:
    return datetime(d.year, d.month, d.day, t.hour, t.minute, t.second)

def list_slots(path: str, sheet_name: str, header_row: int) -> List[Slot]:
    ensure_tracking_columns(path, sheet_name, header_row)
    wb = load_wb(path)
    ws = wb[sheet_name]
    headers = get_headers(ws, header_row)

    def col(name: str) -> int:
        return headers[name.upper()]

    slots: List[Slot] = []
    for r in range(header_row + 1, ws.max_row + 1):
        event = ws.cell(r, col("EVENT")).value
        dval = ws.cell(r, col("DATE")).value
        stval = ws.cell(r, col("START TIME")).value
        enval = ws.cell(r, col("END TIME")).value
        if not (event or dval or stval or enval):
            continue
        event_s = (str(event).strip() if event else "")
        if not event_s:
            continue
        d = parse_excel_date(dval)
        stt = parse_excel_time(stval)
        ent = parse_excel_time(enval)
        if not (d and stt and ent):
            continue
        start_dt = slot_datetime(d, stt)
        end_dt = slot_datetime(d, ent)
        if end_dt <= start_dt:
            end_dt = end_dt + timedelta(days=1)

        slots.append(
            Slot(
                row=r,
                event=event_s,
                location=str(ws.cell(r, col("LOCATION")).value or "").strip(),
                start_dt=start_dt,
                end_dt=end_dt,
                hours=float(ws.cell(r, col("HOURS")).value or 0),
                contact=str(ws.cell(r, col("CONTACT PERSON")).value or "").strip(),
                first_name=str(ws.cell(r, col("FIRST NAME")).value or "").strip(),
                last_name=str(ws.cell(r, col("LAST NAME")).value or "").strip(),
                completed=str(ws.cell(r, col("COMPLETED")).value or "").strip(),
                email=str(ws.cell(r, col("EMAIL")).value or "").strip(),
            )
        )
    return slots

def reserve_excel_slot(path: str, sheet_name: str, header_row: int, row: int, user: Dict) -> Tuple[bool, str]:
    ensure_tracking_columns(path, sheet_name, header_row)
    wb = load_wb(path)
    ws = wb[sheet_name]
    headers = get_headers(ws, header_row)
    def col(n: str) -> int: return headers[n.upper()]

    fn_cell = ws.cell(row, col("FIRST NAME"))
    ln_cell = ws.cell(row, col("LAST NAME"))
    if str(fn_cell.value or "").strip() or str(ln_cell.value or "").strip():
        wb.close()
        return False, "That slot is already taken."

    fn_cell.value = user["first_name"].strip()
    ln_cell.value = user["last_name"].strip()
    ws.cell(row, col("EMAIL")).value = user["email"].strip().lower()
    ws.cell(row, col("USER_ID")).value = user["id"]
    ws.cell(row, col("RESERVED_AT")).value = datetime.now()
    ws.cell(row, col("SENT_24H")).value = False
    ws.cell(row, col("SENT_1H")).value = False
    ws.cell(row, col("GCAL_UID")).value = ""

    wb.save(path)
    return True, "Reserved!"

def cancel_excel_reservation(path: str, sheet_name: str, header_row: int, row: int, email: str) -> Tuple[bool, str]:
    ensure_tracking_columns(path, sheet_name, header_row)
    wb = load_wb(path)
    ws = wb[sheet_name]
    headers = get_headers(ws, header_row)
    def col(n: str) -> int: return headers[n.upper()]

    slot_email = str(ws.cell(row, col("EMAIL")).value or "").strip().lower()
    if slot_email != email.strip().lower():
        wb.close()
        return False, "You can only cancel your own reservation."
    completed = str(ws.cell(row, col("COMPLETED")).value or "").strip()
    if completed:
        wb.close()
        return False, "This slot is marked completed. Contact an admin."

    ws.cell(row, col("FIRST NAME")).value = ""
    ws.cell(row, col("LAST NAME")).value = ""
    ws.cell(row, col("EMAIL")).value = ""
    ws.cell(row, col("USER_ID")).value = ""
    ws.cell(row, col("RESERVED_AT")).value = ""
    ws.cell(row, col("SENT_24H")).value = False
    ws.cell(row, col("SENT_1H")).value = False
    ws.cell(row, col("GCAL_UID")).value = ""

    wb.save(path)
    return True, "Cancelled."

def admin_add_slot_excel(
    path: str, sheet_name: str, header_row: int,
    event: str, location: str, d: date, start_t: time, end_t: time,
    hours: float, contact: str
) -> Tuple[bool, str]:
    ensure_tracking_columns(path, sheet_name, header_row)
    wb = load_wb(path)
    ws = wb[sheet_name]
    headers = get_headers(ws, header_row)
    def col(n: str) -> int: return headers[n.upper()]

    r = header_row + 1
    while r <= ws.max_row:
        ev = ws.cell(r, col("EVENT")).value
        dv = ws.cell(r, col("DATE")).value
        stv = ws.cell(r, col("START TIME")).value
        if not (ev or dv or stv):
            break
        r += 1
    if r > ws.max_row:
        r = ws.max_row + 1

    ws.cell(r, col("EVENT")).value = event.strip()
    ws.cell(r, col("LOCATION")).value = location.strip()
    ws.cell(r, col("DATE")).value = datetime(d.year, d.month, d.day)
    ws.cell(r, col("START TIME")).value = start_t
    ws.cell(r, col("END TIME")).value = end_t
    ws.cell(r, col("HOURS")).value = float(hours)
    ws.cell(r, col("CONTACT PERSON")).value = contact.strip()

    ws.cell(r, col("FIRST NAME")).value = ""
    ws.cell(r, col("LAST NAME")).value = ""
    ws.cell(r, col("COMPLETED")).value = ""
    ws.cell(r, col("EMAIL")).value = ""
    ws.cell(r, col("USER_ID")).value = ""
    ws.cell(r, col("RESERVED_AT")).value = ""
    ws.cell(r, col("SENT_24H")).value = False
    ws.cell(r, col("SENT_1H")).value = False
    ws.cell(r, col("GCAL_UID")).value = ""

    wb.save(path)
    return True, f"Added slot at row {r}."


# ----------------------------
# Email + Calendar
# ----------------------------
def fmt_ics_dt(dt: datetime) -> str:
    return dt.strftime("%Y%m%dT%H%M%S")

def make_ics(event_title: str, start_dt: datetime, end_dt: datetime, location: str, description: str, uid: str) -> str:
    return "\n".join([
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//CUSA//EN",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        "BEGIN:VEVENT",
        f"UID:{uid}",
        f"DTSTAMP:{fmt_ics_dt(datetime.now())}",
        f"DTSTART:{fmt_ics_dt(start_dt)}",
        f"DTEND:{fmt_ics_dt(end_dt)}",
        f"SUMMARY:{event_title}",
        f"LOCATION:{location}",
        f"DESCRIPTION:{description}",
        "END:VEVENT",
        "END:VCALENDAR",
        ""
    ])

def google_calendar_link(title: str, start_dt: datetime, end_dt: datetime, location: str, details: str) -> str:
    dates = f"{start_dt.strftime('%Y%m%dT%H%M%S')}/{end_dt.strftime('%Y%m%dT%H%M%S')}"
    return (
        "https://calendar.google.com/calendar/render?action=TEMPLATE"
        f"&text={quote(title)}"
        f"&dates={quote(dates)}"
        f"&location={quote(location)}"
        f"&details={quote(details)}"
    )

def send_email_with_ics(to_email: str, subject: str, body_text: str, ics_text: str) -> None:
    # If SMTP isn't configured, raise so UI can tell you clearly.
    if not smtp_ready():
        raise RuntimeError("SMTP not configured. Fill .streamlit/secrets.toml (SMTP_HOST/SMTP_USER/SMTP_PASS/FROM_EMAIL) and restart.")
    msg = EmailMessage()
    msg["From"] = FROM_EMAIL
    msg["To"] = to_email
    msg["Subject"] = subject
    msg.set_content(body_text)
    msg.add_attachment(ics_text.encode("utf-8"), maintype="text", subtype="calendar", filename="event.ics")

    with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=30) as s:
        s.starttls()
        s.login(SMTP_USER, SMTP_PASS)
        s.send_message(msg)


# ----------------------------
# App State
# ----------------------------
if "user" not in st.session_state:
    st.session_state.user = None

def do_logout():
    st.session_state.user = None
    st.rerun()


# ----------------------------
# Login / Signup
# ----------------------------
if not st.session_state.user:
    tabs = st.tabs(["Login", "Sign up"])

    with tabs[0]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Login")
        email = st.text_input("Email", key="login_email")
        pw = st.text_input("Password", type="password", key="login_pw")
        if st.button("Login", type="primary", use_container_width=True):
            u = authenticate(email, pw)
            if u:
                if u["email"].strip().lower() in FORCED_ADMIN_EMAILS:
                    set_user_admin(u["email"], True)
                    u["is_admin"] = True
                st.session_state.user = u
                st.success("Logged in!")
                st.rerun()
            else:
                st.error("Invalid email or password.")
        st.markdown("</div>", unsafe_allow_html=True)

    with tabs[1]:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Create account")
        fn = st.text_input("First name", key="su_fn")
        ln = st.text_input("Last name", key="su_ln")
        email2 = st.text_input("Email", key="su_email")
        pw2 = st.text_input("Password (min 8 chars)", type="password", key="su_pw")
        if st.button("Create account", use_container_width=True):
            is_admin = email2.strip().lower() in FORCED_ADMIN_EMAILS
            ok, msg = create_user(fn, ln, email2, pw2, is_admin=is_admin)
            st.success(msg + " Please login.") if ok else st.error(msg)
        st.markdown("</div>", unsafe_allow_html=True)

    st.stop()


# ----------------------------
# Logged-in view
# ----------------------------
u = st.session_state.user

top = st.columns([1, 5, 2])
with top[0]:
    p = load_profile_pic(u["email"])
    if p:
        st.image(p, width=56)
with top[1]:
    st.markdown(f"**{u['first_name']} {u['last_name']}**  ·  `{u['email']}`")
    st.markdown(f"<span class='pill'>{'Admin' if u.get('is_admin') else 'User'}</span>", unsafe_allow_html=True)
with top[2]:
    if st.button("Log out", use_container_width=True):
        do_logout()

# Excel load
try:
    ensure_tracking_columns(EXCEL_PATH, SHEET_NAME, HEADER_ROW)
    slots = list_slots(EXCEL_PATH, SHEET_NAME, HEADER_ROW)
except Exception as e:
    st.error(f"Excel load failed: {e}")
    st.stop()

tabs = st.tabs(["✅ Signups", "💬 Forum", "👤 Profile", "🛠 Admin"])


# ----------------------------
# Signups tab
# ----------------------------
with tabs[0]:
    st.subheader("Signups")

    c1, c2 = st.columns(2)
    with c1:
        upcoming_only = st.toggle("Upcoming only", value=True)
    with c2:
        show_taken = st.toggle("Show taken slots", value=False)

    search = st.text_input("Search (event/location/contact)", placeholder="e.g. Open House, Commons, Kevin...")

    now = datetime.now()
    filtered: List[Slot] = []
    for s in slots:
        is_taken = bool(s.first_name.strip() or s.last_name.strip())
        if upcoming_only and s.start_dt < now:
            continue
        if (not show_taken) and is_taken:
            continue
        if search:
            hay = f"{s.event} {s.location} {s.contact}".lower()
            if search.lower() not in hay:
                continue
        filtered.append(s)

    filtered.sort(key=lambda x: x.start_dt)

    my_res = [s for s in slots if s.email.strip().lower() == u["email"].lower()]
    my_res.sort(key=lambda x: x.start_dt)

    with st.expander("✅ My Reservations", expanded=True):
        if not my_res:
            st.info("You don’t have any reservations yet.")
        else:
            for s in my_res:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                st.markdown(f"**{s.event}**")
                st.markdown(
                    f"<span class='pill'>{s.start_dt.strftime('%a %b %d, %Y %I:%M %p')}</span>"
                    f"<span class='pill'>{s.location or 'No location'}</span>",
                    unsafe_allow_html=True
                )
                st.markdown(f"<span class='pill'>Contact: {s.contact or '—'}</span>", unsafe_allow_html=True)

                b1, b2 = st.columns(2)
                with b1:
                    if st.button("Cancel", key=f"cancel_{s.row}", use_container_width=True):
                        ok, msg = cancel_excel_reservation(EXCEL_PATH, SHEET_NAME, HEADER_ROW, s.row, u["email"])
                        st.success(msg) if ok else st.error(msg)
                        st.rerun()
                with b2:
                    st.caption("To edit: reserve a new slot, then cancel the old one.")
                st.markdown("</div>", unsafe_allow_html=True)

    st.markdown("### Available Slots")
    if not filtered:
        st.info("No matching slots found.")
    else:
        def label(s: Slot) -> str:
            taken = "✅ Taken" if (s.first_name.strip() or s.last_name.strip()) else "🟢 Open"
            return f"{taken} · {s.start_dt:%b %d %I:%M %p} · {s.event} · {s.location}"

        pick = st.selectbox("Choose a slot", filtered, format_func=label)

        if st.button("Reserve this slot", type="primary", use_container_width=True):
            ok, msg = reserve_excel_slot(EXCEL_PATH, SHEET_NAME, HEADER_ROW, pick.row, u)
            if not ok:
                st.error(msg)
                st.stop()

            title = f"{pick.event} (Student Ambassador)"
            details = f"Signed up via CUSA portal. Contact: {pick.contact}"
            gcal = google_calendar_link(title, pick.start_dt, pick.end_dt, pick.location, details)
            ics_uid = str(uuid.uuid4())
            ics = make_ics(title, pick.start_dt, pick.end_dt, pick.location, details, ics_uid)

            body = (
                f"Hi {u['first_name']},\n\n"
                f"✅ You're confirmed for:\n"
                f"Event: {pick.event}\n"
                f"Location: {pick.location}\n"
                f"When: {pick.start_dt:%A, %B %d, %Y at %I:%M %p}\n\n"
                f"Add to Google Calendar:\n{gcal}\n\n"
                f"You will receive reminders 24 hours and 1 hour before.\n\n"
                f"— Cleary University Student Ambassadors"
            )

            try:
                send_email_with_ics(
                    to_email=u["email"],
                    subject=f"Confirmation: {pick.event} ({pick.start_dt:%b %d %I:%M %p})",
                    body_text=body,
                    ics_text=ics
                )
                st.success("Reserved! Confirmation email sent (with calendar invite).")
            except Exception as e:
                st.error(f"Reserved, but email FAILED: {e}")

            st.rerun()

    # Admin: SMTP test button so you can verify right away
    if u.get("is_admin"):
        st.markdown("---")
        st.markdown("### Email test (Admin)")
        if st.button("Send test email to me", use_container_width=True):
            try:
                send_email_with_ics(
                    to_email=u["email"],
                    subject="CUSA Test Email",
                    body_text="This is a test email from the CUSA app. If you received this, SMTP is configured correctly.",
                    ics_text=make_ics("CUSA Test", datetime.now()+timedelta(minutes=5), datetime.now()+timedelta(minutes=35), "Cleary", "SMTP test", str(uuid.uuid4()))
                )
                st.success("Test email sent!")
            except Exception as e:
                st.error(str(e))


# ----------------------------
# Forum tab (scrollable feed + pics + moderation)
# ----------------------------
with tabs[1]:
    st.subheader("Forum")

    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown("### Ask a question")
    f_title = st.text_input("Title", placeholder="e.g., Where do we check in?", key="forum_title")
    f_body = st.text_area("Question details", placeholder="Add context, date/event name, etc.", height=120, key="forum_body")
    if st.button("Post Question", type="primary", use_container_width=True):
        if not f_title.strip() or not f_body.strip():
            st.error("Please add a title and details.")
        else:
            author_name = f"{u['first_name']} {u['last_name']}".strip()
            create_post(u["email"], author_name, f_title, f_body)
            st.success("Posted!")
            st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)

    include_deleted = False
    if u.get("is_admin"):
        include_deleted = st.toggle("Admin: show deleted posts", value=False)

    posts = list_posts(limit=100, include_deleted=include_deleted)

    if not posts:
        st.info("No posts yet. Be the first to ask!")
    else:
        feed = st.container(height=520, border=True)
        with feed:
            for (pid, created_at, author_email, author_name, title, body, locked, deleted) in posts:
                st.markdown('<div class="card">', unsafe_allow_html=True)
                row = st.columns([1, 6, 2])
                with row[0]:
                    pic = load_profile_pic(author_email)
                    if pic:
                        st.image(pic, width=44)
                with row[1]:
                    st.markdown(f"**{title}**")
                    status_bits = []
                    if locked: status_bits.append("🔒 locked")
                    if deleted: status_bits.append("🗑 deleted")
                    status_txt = (" · " + ", ".join(status_bits)) if status_bits else ""
                    st.markdown(
                        f"<span class='pill'>{author_name}</span>"
                        f"<span class='pill'>{created_at}</span>"
                        f"<span class='pill'>#{pid}</span>"
                        f"<span class='pill'>{status_txt[3:] if status_txt else 'active'}</span>",
                        unsafe_allow_html=True
                    )
                    st.write(body)
                with row[2]:
                    if u.get("is_admin"):
                        if st.button("Lock" if not locked else "Unlock", key=f"lock_{pid}", use_container_width=True):
                            set_lock(pid, locked=not bool(locked))
                            st.rerun()
                        if st.button("Delete", key=f"del_{pid}", use_container_width=True, disabled=bool(deleted)):
                            soft_delete_post(pid)
                            st.rerun()

                exp = st.expander("Replies / Reply", expanded=False)
                with exp:
                    reps = list_replies(pid, include_deleted=u.get("is_admin", False))
                    if reps:
                        for rid, r_created, r_email, r_author, r_body, r_deleted in reps:
                            st.markdown('<div class="card">', unsafe_allow_html=True)
                            rr = st.columns([1, 6, 2])
                            with rr[0]:
                                rpic = load_profile_pic(r_email)
                                if rpic:
                                    st.image(rpic, width=36)
                            with rr[1]:
                                st.markdown(
                                    f"<span class='pill'>{r_author}</span><span class='pill'>{r_created}</span>"
                                    + ("<span class='pill'>🗑 deleted</span>" if r_deleted else ""),
                                    unsafe_allow_html=True
                                )
                                st.write(r_body)
                            with rr[2]:
                                if u.get("is_admin"):
                                    if st.button("Delete reply", key=f"delr_{rid}", use_container_width=True, disabled=bool(r_deleted)):
                                        soft_delete_reply(rid)
                                        st.rerun()
                            st.markdown("</div>", unsafe_allow_html=True)
                    else:
                        st.info("No replies yet.")

                    if deleted:
                        st.warning("This post is deleted. Replies are disabled.")
                    elif locked:
                        st.warning("This thread is locked. Replies are disabled.")
                    else:
                        reply = st.text_area("Write a reply", height=90, key=f"reply_box_{pid}", placeholder="Be respectful and helpful 🙂")
                        if st.button("Reply", key=f"reply_btn_{pid}", use_container_width=True):
                            if reply.strip():
                                try:
                                    create_reply(pid, u["email"], f"{u['first_name']} {u['last_name']}".strip(), reply)
                                    st.rerun()
                                except Exception as e:
                                    st.error(str(e))
                            else:
                                st.error("Reply cannot be empty.")
                st.markdown("</div>", unsafe_allow_html=True)


# ----------------------------
# Profile tab
# ----------------------------
with tabs[2]:
    st.subheader("Profile")
    st.markdown('<div class="card">', unsafe_allow_html=True)
    st.markdown(f"**Name:** {u['first_name']} {u['last_name']}")
    st.markdown(f"**Email:** {u['email']}")
    st.markdown(f"<span class='pill'>{'Admin' if u.get('is_admin') else 'User'}</span>", unsafe_allow_html=True)

    current = load_profile_pic(u["email"])
    if current:
        st.image(current, width=160)

    upload = st.file_uploader("Upload profile picture (PNG/JPG)", type=["png", "jpg", "jpeg"])
    if upload is not None:
        save_profile_pic(u["email"], upload.getvalue())
        st.success("Profile picture updated!")
        st.rerun()
    st.markdown("</div>", unsafe_allow_html=True)


# ----------------------------
# Admin tab
# ----------------------------
with tabs[3]:
    if not u.get("is_admin"):
        st.info("Admin tools are restricted.")
    else:
        st.subheader("Admin")

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Add a new slot (adds a new row to Excel)")
        ev = st.text_input("Event name")
        loc = st.text_input("Location")
        d = st.date_input("Date", value=(datetime.now() + timedelta(days=7)).date())
        c1, c2 = st.columns(2)
        with c1:
            stt = st.time_input("Start time", value=datetime.now().replace(minute=0, second=0, microsecond=0).time())
        with c2:
            ent = st.time_input("End time", value=(datetime.now() + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0).time())
        hrs = st.number_input("Hours", min_value=0.0, value=1.0, step=0.5)
        contact = st.text_input("Contact person")

        if st.button("Add slot to Excel", type="primary", use_container_width=True):
            if not ev.strip():
                st.error("Event name is required.")
            else:
                ok, msg = admin_add_slot_excel(EXCEL_PATH, SHEET_NAME, HEADER_ROW, ev, loc, d, stt, ent, hrs, contact)
                st.success(msg) if ok else st.error(msg)
                st.rerun()
        st.markdown("</div>", unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### Admin management (promote other admins)")
        st.caption("User must sign up first, then you can promote their email here.")
        new_admin_email = st.text_input("User email to promote", placeholder="someone@my.cleary.edu")
        if st.button("Make admin", type="primary", use_container_width=True):
            ok, msg = set_user_admin(new_admin_email, True)
            st.success(msg) if ok else st.error(msg)

        remove_admin_email = st.text_input("Admin email to demote", placeholder="someone@my.cleary.edu", key="demote_email")
        if st.button("Remove admin", use_container_width=True):
            if remove_admin_email.strip().lower() in FORCED_ADMIN_EMAILS:
                st.error("That admin is locked as forced-admin in code.")
            else:
                ok, msg = set_user_admin(remove_admin_email, False)
                st.success(msg) if ok else st.error(msg)
        st.markdown("</div>", unsafe_allow_html=True)
